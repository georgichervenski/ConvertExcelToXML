"""
Extracting indicators and their coordinates from excel files
"""
import openpyxl
import pprint
wb = ''
while True:
    try:
        # inputFileToRead = 'C:\\Users\\georgi.chervenski\\Desktop\BNB_Files\\v13_Excel_Templates_Without_Sum\\FRAUD_TEST.xlsx'
        inputFileToRead = input('Please enter the full path to the file with indicators: ')
        wb = openpyxl.load_workbook(inputFileToRead, data_only=True)
    except FileNotFoundError:
        print('Could not find the file! Please try again!')
        continue

    else:
        print('The file path is correct.')
        break

sheetDict = {}

#Loop through all sheetnames and store them in a dict
for every_sheet in wb.sheetnames:
    sheetDict.update({every_sheet: {}})

for key, value in sheetDict.items():
    sheet = wb.get_sheet_by_name(key)
    for row in sheet:
        for cell in row:
            if type(cell.value) == str and cell.value is not None and cell.value.startswith('{') and cell.value.endswith('}'):
                value.update({str(cell.value)[1:-1]: str(cell.coordinate)})
#pprint.pprint(sheetDict)




"""
for every_sheet in wb.sheetnames:
    sheetDict.append(every_sheet)
#Loop through every cell and fill its coordinate and value in a dictionary
for s in iter(sheetDict):
    sheet = wb.get_sheet_by_name(s)
    for row in sheet:
        for cell in row:
            cellsDict.update({str(cell.value): str(cell.coordinate)})
            #cellsDict[s] = {str(cell.value): str(cell.coordinate)}
del cellsDict['None']

#Populate a dictionary with indicators and their coordinates. The curly brackets are removed!
for k, v in cellsDict.items():
    if k.startswith('{') and k.endswith('}'):
        indicatorsDict.update({k[1:-1]: v})
pprint.pprint(cellsDict)
"""