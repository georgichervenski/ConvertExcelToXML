"""
Enter full path to the source file in ExtractDataFromExcel.py
Enter full path to Savka`s File
Enter the xml file name
"""
from openpyxl import load_workbook
from ConvertExcelToXML.Released_version import CreateFileXML as CR, ExtractDataFromExcel as EE
from pprint import pprint
wb = ''
while True:
    try:
        # Please enter the full path to the destination file
        # inputFileToRead = 'C:\\Users\\georgi.chervenski\\Desktop\\BNB_Files\\Files From Savka\\FRAUD.xlsx'
        inputFileToRead = input('Please enter the full path to the data file: ')
        wb = load_workbook(inputFileToRead, data_only=True)
    except FileNotFoundError:
        print('Could not find the file! Please try again!')
        continue

    else:
        print('The file path is correct.')
        break
# Please choose xml file name
inputDestinationXMLFileName = input('Please enter the xml file name: ')

# Create dicts and lists to store data
sheetList = {}
excelFileResults = EE.sheetDict

# Loop through all sheet names and store them in a dict
for every_sheet in wb.sheetnames:
    sheetList.update({every_sheet: {}})

# Loop through every cell in every sheet and fill its coordinate and value in a dictionary
for key, value in sheetList.items():
    sheet = wb.get_sheet_by_name(key)
    for row in sheet:
        for cell in row:
            value.update({str(cell.coordinate): str(cell.value)})

# Map every indicator with its value and store it in a dict
for (k1, v1), (k2, v2) in zip(sheetList.items(), excelFileResults.items()):
    if k1 == k2:
        for kk2, vv2 in dict(v2).items():
            for kk1, vv1 in dict(v1).items():
                if vv2 == kk1:
                    excelFileResults[k1].update({kk2: vv1})
# Call functions which create xml file
CR.add_tr_td(excelFileResults)
CR.createXML(inputDestinationXMLFileName)

